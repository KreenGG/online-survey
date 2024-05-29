from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side

from core.apps.surveys.models import Answer, Survey
from core.apps.surveys.services import SurveyService
from core.project import settings

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def fill_choices_charts(ws, survey: Survey) -> None:
    service = SurveyService()
    questions = service.get_choices_questions(survey)
    count = service.get_participants_number(survey)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15

    ws.append([survey.title])
    ws[f"A{ws.max_row}"].style = "20 % - Accent1"
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=22)
    ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")

    ws.append([f"Всего участников: {count}"])
    ws[f"A{ws.max_row}"].style = "20 % - Accent1"
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=16)

    ws.append(["⠀"])

    for question in questions:
        question_title = question.title

        ws.append([question_title])
        ws[f"A{ws.max_row}"].style = "20 % - Accent1"
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=16)
        ws[f"A{ws.max_row}"].border = thin_border
        ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")

        ws.append(["Ответ", "Количество"])
        ws[f"A{ws.max_row}"].style = "20 % - Accent1"
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
        ws[f"A{ws.max_row}"].border = thin_border

        ws[f"B{ws.max_row}"].style = "20 % - Accent1"
        ws[f"B{ws.max_row}"].font = Font(bold=True, size=12)
        ws[f"B{ws.max_row}"].border = thin_border

        choices = question.choices.all()
        for choice in choices:
            count = Answer.objects.filter(question=question, value=choice.title).count()
            ws.append([choice.title, count])
            ws[f"A{ws.max_row}"].border = thin_border
            ws[f"B{ws.max_row}"].border = thin_border

        ws.append(["⠀"])

        labels = Reference(ws, min_col=1, min_row=ws.max_row-len(choices), max_row=ws.max_row-1)
        data = Reference(ws, min_col=2, min_row=ws.max_row-len(choices), max_row=ws.max_row-1)

        chart = PieChart()

        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = question_title
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        chart_location = f"A{ws.max_row}"
        ws.add_chart(chart, chart_location)

        chart.width = 12
        chart.height = 6

        # Отступ между данными графиков
        ws.insert_rows(ws.max_row, amount=12)


def fill_rating_questions(ws, survey: Survey) -> None:
    service = SurveyService()
    questions = service.get_rating_questions(survey)

    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 20

    ws["G4"] = "Средний рейтинг вопросов"
    ws["G4"].style = "20 % - Accent1"
    ws["G4"].font = Font(bold=True, size=22)
    ws["G4"].border = thin_border
    ws.merge_cells("G4:H4")

    ws["G5"] = "Вопрос"
    ws["G5"].style = "20 % - Accent1"
    ws["G5"].font = Font(bold=True, size=12)
    ws["G5"].border = thin_border

    ws["H5"] = "Средний рейтинг"
    ws["H5"].style = "20 % - Accent1"
    ws["H5"].font = Font(bold=True, size=12)
    ws["H5"].border = thin_border

    start_index = 6
    for question in questions:
        ws[f"G{start_index}"] = question["question"].title
        ws[f"G{start_index}"].border = thin_border
        ws[f"H{start_index}"] = question["avg_rating"]
        ws[f"H{start_index}"].border = thin_border

        start_index += 1


def create_report(survey: Survey) -> None:
    wb = Workbook()
    ws = wb.active

    fill_choices_charts(ws, survey)
    fill_rating_questions(ws, survey)

    path = settings.MEDIA_URL + f"{survey.pk}.xlsx"
    wb.save(path)
